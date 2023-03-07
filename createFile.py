import os, sys
import openpyxl
import logging, traceback

sys.path.append("customLibraries")
import Config
import SystemConfig
import subprocess

def getMaxColumn(sheetObj):
    #custom method since python internal implementation is buggy
    col=0
    while True:
        col+=1
        cellValue=trimValue(getCellValue(sheetObj,Config.rowNumberWhereFieldsArePlaced,col))
        print("In getMaxColumn(), cell value:",cellValue)
        if cellValue in ['None', "",None]:
            return col-1

def trimValue(value):
    try:
        return str(value).strip()
    except:
        print "Exception while trimming Cell Value"
        return None

def findStringLocationInSheet(sheet_obj,maxRows,maxColumns,expectedString):
    """
        Handle condition : what if the field is not found, should terminate execution with FATAL
    """
    for currentRow in range(1, maxRows+1):
        for currentCol in range(1,maxColumns+1):
            cell_obj = sheet_obj.cell(row=currentRow, column=currentCol)
            try:
                if str(cell_obj.value).strip()==str(expectedString):
                    return (currentRow,currentCol)
            except Exception,e:
                print traceback.print_exc()

    print "Terminating Execution since Excel Template was tampered. Expected String : {0} was not found".format(expectedString)
    sys.exit(-1)

def basicOperations(excelFile):
    print excelFile
    wb_obj = openpyxl.load_workbook(excelFile, data_only=True)
    sheet_obj = wb_obj[Config.sheetName]
    print "Sheet title :",sheet_obj.title
    maxRows=sheet_obj.max_row

    #including the last column
    maxColumns=getMaxColumn(sheet_obj)

    return (wb_obj,sheet_obj,maxRows,maxColumns)

def getDelimiterFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.DelimiterWhileFileCreation)
    delim= (sheet_obj.cell(row=row, column=col+1)).value
    try:
        delim=delim.replace("[","").replace("]","")
        return delim
    except Exception,e:
        print "Unable to fetch Delimiter from Excel : Terminating Execution "
        traceback.print_exc()
        sys.exit(-1)

def addPadding(delimiter,stringToPad, columnWidth):
    if stringToPad is None:
        stringToPad=""
    return str(stringToPad).ljust(columnWidth,str(delimiter))

def createFile(sheet_obj, maxRows, maxColumns,dictColumnWidth,fileName,delimiter):
    #logic : get the row, col from where you need to start
    #col# will give you the dict count

    #create source file anyway
    #if the file needs to be zipped, zip it with/without password

    (row, col) = getStartingRowAndColumnForFileCreationTags(sheet_obj, maxRows, maxColumns)
    row+=1

    file=open(fileName,"w")

    # field len for each field should be stored in a dict, for faster processing
    for currentRow in range(row,maxRows+1):
        currentStr=""
        for currentCol in range(col,maxColumns+1):
            cellValue = getCellValue(sheet_obj,currentRow,currentCol)
            currentStr+=addPadding(delimiter, cellValue, int(dictColumnWidth[currentCol]))
        file.write(currentStr+"\n")
    file.close()

    #analyze whether the file needs to be zipped or not
    if SystemConfig.zipFile is not None and SystemConfig.zipFile != "":
        #zip the above file

        targetZipFileAbsPath=os.path.join(SystemConfig.sourceFilePath, SystemConfig.zipFile)

        #check if password is needed
        if SystemConfig.sourceFilePassword is None or SystemConfig.sourceFilePassword=="":
            #zip without password
            commandToRun="{0}\\7za a {1} {2}".format(Config.assetsPath,targetZipFileAbsPath,fileName)
        else:
            #zip with password
            commandToRun='{0}\\7za a {1} {2} -p"{3}"'.format(Config.assetsPath, targetZipFileAbsPath,
                                                             fileName, SystemConfig.sourceFilePassword)
        try:
            print("Current dir: ",os.getcwd())
            print("Command to run: ",commandToRun)
            subprocess.check_call(commandToRun)
        except subprocess.CalledProcessError:
            # There was an error - command exited with non-zero code
            print("\n\n\n")
            print("*"*20)
            print("Exception while zipping file, can not continue further")
            print("*"*20)

def getFileNameFromExcel(sheet_obj, maxRows, maxColumns):
    """
        The next column in the matched row will be returned
        This function also sets the source zip file password and zip file name
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.Souce_File_Name)
    fileName= (sheet_obj.cell(row=row, column=col+1)).value.strip()


    #this is the zip file password
    try:
        (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.ZipFileName)
        SystemConfig.sourceFilePassword=(sheet_obj.cell(row=row, column=col+2)).value.strip()
        SystemConfig.sourceFilePassword=str(SystemConfig.sourceFilePassword).replace("ZipPass:","")
        SystemConfig.sourceFilePassword=str(SystemConfig.sourceFilePassword).strip()
    except:
        pass
    #this is the zip file name
    try:
        (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.ZipFileName)
        SystemConfig.zipFile=(sheet_obj.cell(row=row, column=col+1)).value.strip()
    except:
        pass

    print "Source File Name : ",fileName
    return fileName

def getStartingRowAndColumnForFileCreationTags(sheet_obj,maxRows,maxColumns):
    """
    Returns row# and col# for the first tag of the file creation template
    In the same row, the cols shall continue till max col, take a note for capturing the last col
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.lastColumnBeforeFileCreationFields)
    return (row,col+1)

def getCellValue(sheet_obj,row,col):
    try:
        return sheet_obj.cell(row=row, column=col).value
    except Exception,e:
        print(e)
        return None

def parseFieldTagsToSeperateNameAndWidth(cellValue):
    try:
        #print "Handling Parsing for Cell Value : ",cellValue
        fieldName=cellValue.split("[")[0].strip()
        fieldWidth=(cellValue.split("[")[1]).split("]")[0].strip()
        return (fieldName, fieldWidth)
    except Exception,e:
        print "Parsing Error : Terminating Execution since Excel Template was tampered / not created properly. Check the cellValue :",cellValue
        sys.exit(-1)

def storeFieldLengthsforAllFieldTags(sheet_obj,maxRows,maxColumns):
    dictColumnWidth={}
    (row,col)=getStartingRowAndColumnForFileCreationTags(sheet_obj,maxRows,maxColumns)
    print "getStartingRowAndColumnForFileCreationTags() : Row : {0},Col : {1}".format(row,col)
    for currentCol in range(col,maxColumns+1):
        currentCellVal=getCellValue(sheet_obj,row,currentCol)
        if currentCellVal is None:
            print "Cell val is None at {0},{1}".format(row,currentCol)
            sys.exit(-1)
        else:
            print "Analysing cell : {0}".format(currentCellVal)
        (colName,colWidth)=parseFieldTagsToSeperateNameAndWidth(currentCellVal)
        dictColumnWidth[currentCol]=colWidth
    return dictColumnWidth

def printDict(dict):
    print "Printing Dictionary contents : ",dict

def run(excelFile):
    print excelFile
    (wb_obj,sheet_obj,maxRows,maxColumns) = basicOperations(excelFile)
    delimiter = getDelimiterFromExcel(sheet_obj,maxRows,maxColumns)
    print "Max Rows : {0}\nMax Cols : {1}".format(maxRows,maxColumns)
    dictColumnWidth=storeFieldLengthsforAllFieldTags(sheet_obj, maxRows, maxColumns)
    printDict(dictColumnWidth)
    SystemConfig.sourceFilePath = os.getcwd()+"\\FilesToUpload"

    if not os.path.exists(SystemConfig.sourceFilePath):
        os.makedirs(SystemConfig.sourceFilePath)

    sourceFileName=SystemConfig.sourceFilePath+"\\"+getFileNameFromExcel(sheet_obj,maxRows,maxColumns)
    createFile(sheet_obj, maxRows, maxColumns,dictColumnWidth,sourceFileName,delimiter)

if __name__ == '__main__':
    run(Config.ExcelPath)