import os, sys
import openpyxl
import logging, traceback
import paramiko, time

sys.path.append("customLibraries")
import sftpCustom
import SystemConfig
import Config

def findStringLocationInSheet(sheet_obj,maxRows,maxColumns,expectedString):
    """
        Handle condition : what if the field is not found, should terminate execution with FATAL
    """
    for currentRow in range(1, maxRows+1):
        for currentCol in range(1,maxColumns+1):
            cell_obj = sheet_obj.cell(row=currentRow, column=currentCol)
            try:
                if str(cell_obj.value).strip()==str(expectedString):
                    return (currentRow,currentCol)
            except Exception,e:
                print traceback.print_exc()

    print "Terminating Execution since Excel Template was tampered."
    sys.exit(-1)

def basicOperations(excelFile):
    wb_obj = openpyxl.load_workbook(excelFile)
    sheet_obj = wb_obj.active
    print "Sheet title :",sheet_obj.title
    maxRows=sheet_obj.max_row
    maxColumns=sheet_obj.max_column

    return (wb_obj,sheet_obj,maxRows,maxColumns)

def getDelimiterFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.DelimiterWhileFileCreation)
    delim= (sheet_obj.cell(row=row, column=col+1)).value
    try:
        delim=delim.replace("[","").replace("]","")
        return delim
    except Exception,e:
        print "Unable to fetch Delimiter from Excel : Terminating Execution "
        traceback.print_exc()
        sys.exit(-1)

def getZipFilenameFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    zipFileName=None
    try:
        (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.ZipFileName)
        zipFileName= (sheet_obj.cell(row=row, column=col+1)).value
        try:
            if zipFileName is None or zipFileName.strip() is None or zipFileName.strip()=="":
                return None
                #raise Exception('Zip file name is not correctly defined in Excel Template.')
        except:
            pass
        return zipFileName
    except Exception,e:
        traceback.print_exc()
        print "\n\nTerminating Execution since Rename File Name could not be fetched from Excel, kindly check if the value is populated in excel"
        sys.exit(-1)

def addPadding(delimiter,stringToPad, columnWidth):
    return stringToPad.ljust(columnWidth,delimiter)

def createFile(sheet_obj, maxRows, maxColumns,dictColumnWidth,fileName,delimiter):
    #logic : get the row, col from where you need to start
    #col# will give you the dict count

    (row, col) = getStartingRowAndColumnForFileCreationTags(sheet_obj, maxRows, maxColumns)
    row+=1

    file=open(fileName,"w")

    # field len for each field should be stored in a dict, for faster processing
    for currentRow in range(row,maxRows+1):
        currentStr=""
        for currentCol in range(col,maxColumns+1):
            currentStr+=addPadding(delimiter,getCellValue(sheet_obj,currentRow,currentCol),int(dictColumnWidth[currentCol]))

        file.write(currentStr+"\n")
    file.close()

def getFileNameFromExcel(sheet_obj, maxRows, maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.Souce_File_Name)
    fileName= (sheet_obj.cell(row=row, column=col+1)).value.strip()
    print "Source File Name : ",fileName
    return fileName

def getStartingRowAndColumnForFileCreationTags(sheet_obj,maxRows,maxColumns):
    """
    Returns row# and col# for the first tag of the file creation template
    In the same row, the cols shall continue till max col, take a note for capturing the last col
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.lastColumnBeforeFileCreationFields)
    return (row,col+1)

def getCellValue(sheet_obj,row,col):
    return sheet_obj.cell(row=row, column=col).value

def parseFieldTagsToSeperateNameAndWidth(cellValue):
    try:
        #print "Handling Parsing for Cell Value : ",cellValue
        fieldName=cellValue.split("[")[0].strip()
        fieldWidth=(cellValue.split("[")[1]).split("]")[0].strip()
        return (fieldName, fieldWidth)
    except Exception,e:
        print "Parsing Error : Terminating Execution since Excel Template was tampered / not created properly. Check the cellValue :",cellValue
        sys.exit(-1)

def storeFieldLengthsforAllFieldTags(sheet_obj,maxRows,maxColumns):
    dictColumnWidth={}
    (row,col)=getStartingRowAndColumnForFileCreationTags(sheet_obj,maxRows,maxColumns)
    #print "getStartingRowAndColumnForFileCreationTags() : Row : {0},Col : {1}".format(row,col)
    for currentCol in range(col,maxColumns+1):
        currentCellVal=getCellValue(sheet_obj,row,currentCol)
        (colName,colWidth)=parseFieldTagsToSeperateNameAndWidth(currentCellVal)
        dictColumnWidth[currentCol]=colWidth

    return dictColumnWidth

def printDict(dict):
    print "Printing Dictionary contents : ",dict

def getDumpLocationFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.Dump_File_Location)
    dumpLocation= (sheet_obj.cell(row=row, column=col+1)).value
    return dumpLocation

def DumpToLocation(sourceFileName,dumpLocationAbsolute):
    retries = 0
    if(not sftpCustom.Connect()):
        sys.exit(-1)
    sftpCustom.uploadFile(sourceFileName, dumpLocationAbsolute)

def createCronFiles():
    print "[ Creating Cron Files ]"
    for eachFile in Config.cronFileNames:
        file=open(eachFile,"w")
        file.close()

def uploadCronFiles():
    print "[ Uploading Cron Files ]"
    #sftpCustom.Connect()
    for eachFile in Config.cronFileNames:
        print "[cronfile] " + eachFile
        sftpCustom.uploadFile(eachFile,Config.cronFileLocation+r'/'+eachFile)

def triggerCronJobs():
    print "[ Triggering Cron Jobs ]"
    createCronFiles()
    uploadCronFiles()

def logDumpTime():
    file=open("lastDumpTime.py","w")
    file.write("lastDumpTime="+str(int(time.time())))
    file.close()

def run(excelFile):
    (wb_obj,sheet_obj,maxRows,maxColumns)=basicOperations(excelFile)
    zipFileName=getZipFilenameFromExcel(sheet_obj,maxRows,maxColumns)
    if zipFileName is None :
        filename = getFileNameFromExcel(sheet_obj,maxRows,maxColumns)
        sourceFileNameAbsolute=os.getcwd()+"\\FilesToUpload\\" + filename
        dumpLocationAbsolute=getDumpLocationFromExcel(sheet_obj,maxRows,maxColumns) +r'/' + filename
    else:
        sourceFileNameAbsolute=os.getcwd() + "\\FilesToUpload\\" + zipFileName
        dumpLocationAbsolute=getDumpLocationFromExcel(sheet_obj,maxRows,maxColumns) + r'/' + zipFileName

    print "Dump location absolute : ",dumpLocationAbsolute
    print "Source file name absolute : ",sourceFileNameAbsolute
    logDumpTime()
    DumpToLocation(sourceFileNameAbsolute,dumpLocationAbsolute)
    triggerCronJobs()
    sftpCustom.closeConnection()

if __name__ == '__main__':
    run(Config.ExcelPath)