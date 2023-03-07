import os, sys
import openpyxl
import logging, traceback

sys.path.append("customLibraries")
import Config
import SystemConfig
from datetime import datetime, timedelta

class adjustDates():
    def __init__(self, excelFile):
        self.excelFile = excelFile
        self.wb_obj = None
        self.sheet_obj = None
        self.maxRows = None
        self.maxColumns = None
        if SystemConfig.dateOverride is "":
            self.date = datetime.now()
            self.date = self.date - timedelta(days=1)
            while self.date.weekday() > 4:
                self.date = self.date - timedelta(days=1)
        else:
            self.date = datetime.strptime(SystemConfig.dateOverride, '%m%d%Y').date()

    def getCellValue(self, row, col):
        try:
            return self.sheet_obj.cell(row, col).value
        except Exception,e:
            return None

    def setCellValue(self, row, col, value):
        try:
            if value  not in ['None', "", None]:
                self.sheet_obj.cell(row, col).value = value
        except Exception,e:
            return None

    def getMaxColumn(self):
        col = 0
        while True:
            col += 1
            cellValue = self.getCellValue(Config.rowNumberWhereFieldsArePlaced, col)
            if str(cellValue).strip() in ['None', "",None]:
                print("[INF] getMaxColumn::Max Column Found", col - 1)
                return col - 1

    def basicOperations(self):
        self.wb_obj = openpyxl.load_workbook(self.excelFile, data_only=True)
        self.sheet_obj = self.wb_obj[Config.sheetName]
        print "Sheet title :",self.sheet_obj.title

        self.maxRows = self.sheet_obj.max_row
        self.maxColumns = self.getMaxColumn()

    def saveWorkBook(self):
        self.wb_obj.save(self.excelFile)

    def findStringLocationInSheet(self, expectedString):
        for currentRow in range(1, self.maxRows + 1):
            for currentCol in range(1, self.maxColumns + 1):
                cell_obj = self.sheet_obj.cell(currentRow, currentCol)
                try:
                    if str(cell_obj.value).strip() == str(expectedString):
                        return (currentRow, currentCol)
                except Exception,e:
                    print traceback.print_exc()

        print "Terminating Execution since Excel Template was tampered. Expected String : {0} was not found".format(expectedString)
        sys.exit(-1)

    def replacePlaceHolders(self, text):
        if text is None or not isinstance(text, basestring):
            return text

        if "MM" in text:
            text = text.replace("MM", self.date.strftime('%m'))
        if "DD" in text:
            text = text.replace("DD", self.date.strftime('%d'))
        if "YYYY" in text:
            text = text.replace("YYYY", self.date.strftime('%Y'))
        if "YY" in text:
            text = text.replace("YY", self.date.strftime('%y'))
        return text

    def getStartandEndData(self):
        (row, col) = self.findStringLocationInSheet(SystemConfig.lastColumnBeforeFileCreationFields)
        return row, col + 2

    def replaceFileNames(self):
        (row, col) = self.findStringLocationInSheet(SystemConfig.Souce_File_Name)
        fileName = self.getCellValue(row, col + 1)
        fileName = self.replacePlaceHolders(fileName)
        self.setCellValue(row, col + 1, fileName)

        (row, col) = self.findStringLocationInSheet(SystemConfig.RenameFileName)
        fileName = self.getCellValue(row, col + 1)
        fileName = self.replacePlaceHolders(fileName)
        self.setCellValue(row, col + 1, fileName)

        (row, col) = self.findStringLocationInSheet(SystemConfig.ZipFileName)
        fileName = self.getCellValue(row, col + 1)
        fileName = self.replacePlaceHolders(fileName)
        self.setCellValue(row, col + 1, fileName)

    def replaceRows(self):
        startRow, endColumn = self.getStartandEndData()
        for currentRow in range(startRow, self.maxRows + 1):
            for currentCol in range(1, endColumn):
                value = self.getCellValue(currentRow, currentCol)
                value = str(self.replacePlaceHolders(value)).strip()
                self.setCellValue(currentRow, currentCol, value)

    def run(self):
        self.basicOperations()
        x =  self.replaceFileNames()
        self.replaceRows()
        self.saveWorkBook()