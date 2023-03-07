import os
import sys

sys.path.append("customLibraries")
print('\n'.join(sys.path))

import sftpCustom
import openpyxl
import Config
import logging, traceback
import SystemConfig
import paramiko, time

import lastDumpTime
import customLibraries.Report as Report

arFileFound=False
usFileFound=False

def findStringLocationInSheet(sheet_obj,maxRows,maxColumns,expectedString):
    """
        Handle condition : what if the field is not found, should terminate execution with FATAL
    """
    for currentRow in range(1, maxRows+1):
        for currentCol in range(1,maxColumns+1):
            cell_obj = sheet_obj.cell(row=currentRow, column=currentCol)
            try:
                if str(cell_obj.value).strip()==str(expectedString):
                    return (currentRow,currentCol)
            except Exception,e:
                print traceback.print_exc()

    print "Expected String : {0} was not found in excel.".format(expectedString)
    print "Max Columns : {0}".format(maxColumns)
    print "Terminating Execution since Excel Template was tampered."
    sys.exit(-1)

def findSubsetStringLocationInSheet(sheet_obj,maxRows,maxColumns,expectedString):

    """
        Handle condition : what if the field is not found, should terminate execution with FATAL
    """
    for currentRow in range(1, maxRows+1):
        for currentCol in range(1,maxColumns+1):
            cell_obj = sheet_obj.cell(row=currentRow, column=currentCol)
            try:
                if str(expectedString) in str(cell_obj.value).strip():
                    return (currentRow,currentCol)
            except Exception,e:
                print traceback.print_exc()

    print "Expected String : {0} was not found in excel.".format(expectedString)
    print "Max Columns : {0}".format(maxColumns)
    print "Terminating Execution since Excel Template was tampered."
    sys.exit(-1)

def getMaxColumn(sheet_obj, maxRows, maxColumns, columnName):
    #last column which is populated for the row header
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, "US File Output Location")
    ctr=col

    #in the var : row find the col number of the last header
    while True:
        try:
            cell_obj = sheet_obj.cell(row=row, column=ctr)
            if (cell_obj.value).strip()=="" or (cell_obj.value).strip() is None:
                break
            ctr+=1
        except:
            break

    #since the valid col number was one before the blank col
    #print "In func getMaxColumn : Max col : {0}".format(ctr-1)
    time.sleep(3)
    return ctr-1

def basicOperations(excelFile):
    wb_obj = openpyxl.load_workbook(excelFile,data_only=True)
    sheet_obj = wb_obj.active
    print "Sheet title :",sheet_obj.title

    maxRows=sheet_obj.max_row
    maxColumns=getMaxColumn(sheet_obj,maxRows,100,SystemConfig.USFileColNumber)

    return (wb_obj,sheet_obj,maxRows,maxColumns)

def getDelimiterFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.DelimiterWhileFileCreation)
    delim= (sheet_obj.cell(row=row, column=col+1)).value
    try:
        delim=delim.replace("[","").replace("]","")
        return delim
    except Exception,e:
        print "Unable to fetch Delimiter from Excel : Terminating Execution "
        traceback.print_exc()
        sys.exit(-1)

def getRenameFilenameFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    try:
        (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.RenameFileName)
        renameFilename= (sheet_obj.cell(row=row, column=col+1)).value

        if renameFilename is None or renameFilename.strip() is None:
            raise Exception('Rename file name is not correctly defined in Excel Template.')

        return renameFilename

    except Exception,e:
        traceback.print_exc()
        print "\n\nTerminating Execution since Rename File Name could not be fetched from Excel, kindly check if the value is populated in excel"
        sys.exit(-1)

def getZipFilenameFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    zipFileName=None
    try:
        (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.ZipFileName)
        zipFileName= (sheet_obj.cell(row=row, column=col+1)).value

        try:
            if zipFileName is None or zipFileName.strip() is None or zipFileName.strip()=="":
                return None
                #raise Exception('Zip file name is not correctly defined in Excel Template.')
        except:
            pass

        return zipFileName

    except Exception,e:
        traceback.print_exc()
        print "\n\nTerminating Execution since Rename File Name could not be fetched from Excel, kindly check if the value is populated in excel"
        sys.exit(-1)

def addPadding(delimiter,stringToPad, columnWidth):
    return stringToPad.ljust(columnWidth,delimiter)

def createFile(sheet_obj, maxRows, maxColumns,dictColumnWidth,fileName,delimiter):
    #logic : get the row, col from where you need to start
    #col# will give you the dict count

    (row, col) = getStartingRowAndColumnForFileCreationTags(sheet_obj, maxRows, maxColumns)
    row+=1

    file=open(fileName,"w")

    # field len for each field should be stored in a dict, for faster processing
    for currentRow in range(row,maxRows+1):
        currentStr=""
        for currentCol in range(col,maxColumns+1):
            currentStr+=addPadding(delimiter,getCellValue(sheet_obj,currentRow,currentCol),int(dictColumnWidth[currentCol]))

        file.write(currentStr+"\n")
    file.close()

def getFileNameFromExcel(sheet_obj, maxRows, maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.Souce_File_Name)
    fileName= (sheet_obj.cell(row=row, column=col+1)).value.strip()
    print "Source File Name : ",fileName
    return fileName

def getStartingRowAndColumnForFileCreationTags(sheet_obj,maxRows,maxColumns):
    """
    Returns row# and col# for the first tag of the file creation template
    In the same row, the cols shall continue till max col, take a note for capturing the last col
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.lastColumnBeforeFileCreationFields)
    return (row,col+1)

def getCellValue(sheet_obj,row,col):
    return sheet_obj.cell(row=row, column=col).value

def parseFieldTagsToSeperateNameAndWidth(cellValue):
    try:
        #print "Handling Parsing for Cell Value : ",cellValue
        fieldName=cellValue.split("[")[0].strip()
        fieldWidth=(cellValue.split("[")[1]).split("]")[0].strip()
        return (fieldName, fieldWidth)
    except Exception,e:
        print "Parsing Error : Terminating Execution since Excel Template was tampered / not created properly. Check the cellValue :",cellValue
        sys.exit(-1)

def storeFieldLengthsforAllFieldTags(sheet_obj,maxRows,maxColumns):
    dictColumnWidth={}
    (row,col)=getStartingRowAndColumnForFileCreationTags(sheet_obj,maxRows,maxColumns)
    #print "getStartingRowAndColumnForFileCreationTags() : Row : {0},Col : {1}".format(row,col)
    for currentCol in range(col,maxColumns+1):
        currentCellVal=getCellValue(sheet_obj,row,currentCol)
        (colName,colWidth)=parseFieldTagsToSeperateNameAndWidth(currentCellVal)
        dictColumnWidth[currentCol]=colWidth
    return dictColumnWidth

def printDict(dict):
    print "Printing Dictionary contents : ",dict

def getDumpLocationFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.Dump_File_Location)
    dumpLocation= (sheet_obj.cell(row=row, column=col+1)).value
    return dumpLocation

#validations need to take place for raw,rename,error,us,ar
def setColumnNumbersForFileValidations():
    (sheet_obj, maxRows, maxColumns)=(SystemConfig.sheet_obj,SystemConfig.maxRows,SystemConfig.maxCol)

    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.ARAmountField)
    SystemConfig.ARAmountColNumber =col

    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.RawFileLocation)
    SystemConfig.RawFileColNumber =col

    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.USFileOutputLocation)
    SystemConfig.USFileColNumber = col

    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.ARFileOutputLocation)
    SystemConfig.ARFileColNumber = col

    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.ErrorFileOutputLocation)
    SystemConfig.ErrorFileColNumber = col

    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.RenameFileLocation)
    SystemConfig.RenameFileColNumber = col

    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.IdentifierInARFile)
    SystemConfig.IdentifierInARFileColNumber = col

    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.USFileDepositoryAccount)
    SystemConfig.USFileDepositoryAccountColNumber = col

    (row, col) = findSubsetStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.depositAccount)
    SystemConfig.depositAccountColNumber = col


def validateRawFile(rowNumber):
    rawFileLocation=getCellValue(SystemConfig.sheet_obj,rowNumber,SystemConfig.RawFileColNumber)

    if(rawFileLocation[-1]=="/"):
        rawFileLocation=rawFileLocation[:-1]

    #raw file name is same as source file name

    if SystemConfig.zipFile is None:
        print "Raw file location : {0}\nSourceFileName: {1}".format(rawFileLocation,SystemConfig.sourceFileName)
        rawFileNameAbs=rawFileLocation+r"/"+SystemConfig.sourceFileName
    else:
        print "Raw file location : {0}\nSourceFileName: {1}".format(rawFileLocation,SystemConfig.zipFile)
        rawFileNameAbs=rawFileLocation+r"/"+SystemConfig.zipFile

    #go to location and check if raw file is created or not
    if sftpCustom.checkIfFileExists(rawFileNameAbs):
        Report.WriteTestStep("Checking if Raw File Exists. Absolute path : {0}".format(rawFileNameAbs),
                             "Raw file should exist","Raw file exists","Passed")
    else:
        Report.WriteTestStep("Checking if Raw File Exists. Absolute path : {0}".format(rawFileLocation),
                             "Raw file should exist", "Raw file does not exist", "Failed")
        print "[ERR] Raw file has not been processed for " + rawFileLocation
        Report.GeneratePDFReport()
        sftpCustom.closeConnection()
        sys.exit(-1)

def validateRenameFile(rowNumber):
    renameFileLocation = getCellValue(SystemConfig.sheet_obj, rowNumber, SystemConfig.RenameFileColNumber)
    renameFileName=getRenameFilenameFromExcel(SystemConfig.sheet_obj,SystemConfig.maxRows,SystemConfig.maxCol)

    if(renameFileLocation[-1]=="/"):
        renameFileLocation=renameFileLocation[:-1]

    renameFileNameAbs=renameFileLocation+r"/"+renameFileName

    #go to location and check if rename file is created or not
    if sftpCustom.checkIfFileExists(renameFileNameAbs):
        Report.WriteTestStep("Checking if Rename File Exists. Absolute path : {0}".format(renameFileNameAbs),
                             "Rename file should exist","Rename file exists","Passed")

    else:
        Report.WriteTestStep("Checking if Rename File Exists. Absolute path : {0}".format(renameFileNameAbs),
                             "Rename file should exist", "Rename file does not exist", "Failed")

def checkIfErrorFileExists(rowNumber):
    errorFileLocation = getCellValue(SystemConfig.sheet_obj, rowNumber, SystemConfig.ErrorFileColNumber)

    print "Error file location : ",errorFileLocation

    if str(errorFileLocation).strip() is None:
        print "No error file location populated"
        return False

    errorFileName = getRenameFilenameFromExcel(SystemConfig.sheet_obj, SystemConfig.maxRows,
                                               SystemConfig.maxCol) + "_Reject"

    if (errorFileLocation[-1] == "/"):
        errorFileLocation = errorFileLocation[:-1]

    errorFileNameAbs = errorFileLocation + r"/" + errorFileName

    errorFileExists = False
    # go to location and check if rename file is created or not
    if sftpCustom.checkIfFileExists(errorFileNameAbs):
        errorFileExists = True
        Report.WriteTestStep("Checking if Error File Exists. Absolute path : {0}".format(errorFileNameAbs),
                             "Error file should exist", "Error file exists", "Passed")
    else:
        Report.WriteTestStep("Checking if Error File Exists. Absolute path : {0}".format(errorFileNameAbs),
                             "Error file should exist", "Error file does not exist", "Failed")

    return (errorFileNameAbs,errorFileExists)

def downloadRemoteErrorFileToLocal(errorFileNameAbs):

    try:
        SystemConfig.errorFileTempDir = r"DownloadedFilesTemp\ErrorFiles"

        if not os.path.exists(SystemConfig.errorFileTempDir):
            os.makedirs(SystemConfig.errorFileTempDir)

        SystemConfig.errorFileTemp = SystemConfig.errorFileTempDir + "\errorFile_" + str(int(time.time()))

    except Exception, e:
        Report.WriteTestStep("Exception occured while creating directory on Local File System",
                             "Should be able to create directory",
                             "Not able to create directory", "Failed")
        traceback.print_exc()
        return False

    # download file
    downloadFileStatus = sftpCustom.downloadFile(errorFileNameAbs, SystemConfig.errorFileTemp)

    if downloadFileStatus:
        Report.WriteTestStep("Downloading of remote file successful : {0}".format(errorFileNameAbs),
                             "Should be able to download file", "Able to download file", "Passed")
    else:
        Report.WriteTestStep("Downloading of remote file failed : {0}".format(errorFileNameAbs),
                             "Should be able to download file", "Unable to download file", "Failed")

    return downloadFileStatus

def validateErrorFile(rowNumber):
    (errorFileNameAbs, errorFileExists)=checkIfErrorFileExists(rowNumber)

    if not errorFileExists:
        Report.WriteTestStep("Error file not found. Content validaiton will be skipped.",
                             "Error file should exist", "Error file does not exist", "Failed")
        return False
    else:
        if downloadRemoteErrorFileToLocal(errorFileNameAbs):
            validateDownloadedErrorFile(rowNumber)

def readFileContentsAsString(fileToRead):
    data=""
    with open(fileToRead, 'r') as file:
        data = data + file.read()

    return data

def readFileContentsAsList(fileToRead):
    data=[]
    with open(fileToRead, 'r') as file:
        data = file.readlines()

    return data

def validateDownloadedErrorFile(rowNumber):
    listIndex=rowNumber-SystemConfig.startingRowNumberForRecordProcessing-1
    print "Row number : {0}".format(listIndex)
    #read contents of raw file, then error file, then check if raw is contained in error file or not
    #sourceFileContents=readFileContentsAsList("FilesToUpload\\"+getFileNameFromExcel(SystemConfig.sheet_obj,SystemConfig.maxRows,SystemConfig.maxCol))

    sourceFileContents=SystemConfig.sourceFileCurrentContent


    print "sourceFileContents as List : {0}".format(sourceFileContents)

    errorFileContents=readFileContentsAsString(SystemConfig.errorFileTemp)


    if sourceFileContents in errorFileContents:

        print "Validation passed"
        Report.WriteTestStep("Error file Content validation Passed","All contents of raw file should be present in Error file","Contents of Raw file are NOT present error file. Error file contents : {0}\n\nSource file contents : {1}".format(errorFileContents,sourceFileContents),"Passed")

    else:
        print "Validation failed"
        Report.WriteTestStep("Error file Content validation Failed",
                             "All contents of raw file should be present in Error file",
                             "Contents of Raw file are NOT present error file. Error file contents : {0}\n\nSource file contents : {1}".format(
                                 errorFileContents,sourceFileContents), "Failed")

def checkArFileContent(arFile):
    #store amount field from excel
    #open each ar file and check if rename file name is contained, if yes
    #check if the contained rename file is the 2nd field of the first row
    #if yes find the row, and match the amount in that specific row
    #its clear that the importance of line here is significant

    #assumption is ar file is delimited by ,

    global arFileFound
    global usFileFound
    stopProcessing=False

    lineList=[]
    #file=open(arFile,"r")


    with open(arFile) as f:
        lineList = f.readlines()



    line=lineList[0]

    renameFileName = getRenameFilenameFromExcel(SystemConfig.sheet_obj,SystemConfig.maxRows,SystemConfig.maxCol)

    print "Searching for Rename file : ",renameFileName
    #time.sleep(5)

    #validation 1
    if "non-migrated" not in SystemConfig.arFileLocation:
        if  renameFileName in line:
            arFileFound=True
            Report.WriteTestStep("Rename filename found in file : {0}".format(arFile),"Contents of the AR File : {0}".format(line),"Same as expected","Passed");
            try:
                secondField = line.split(SystemConfig.delimiterInArFile)[1]

                # if renameFileName.strip() in secondField.strip():
                #     Report.WriteTestStep("AR File Validation : 2nd field should be name of rename file : {0}".format(
                #         renameFileName), "Second field should be name of rename file",
                #                      "Second field is the name of rename file", "Passed")
                    #return True


            except Exception, e:
                # parsing failure
                # log it in PDF
                traceback.print_exc()
                Report.WriteTestStep("AR File Validation : 2nd field should be name of rename file : {0}".format(
                    renameFileName), "Second field should be the name of rename file",
                    "Could not identify 2nd field in the AR File Header. Please check if delimiter is : {0}".format(
                        SystemConfig.delimiterInArFile), "Failed")

                return (False,stopProcessing)
        else:
            return  (False,stopProcessing)

    #if the AR file identifier is contained in the file

    print "\nSystemConfig.IdentifierInARFile:",SystemConfig.IdentifierInARFile
    print "\nLine List is : ",lineList
    identifierFoundInArFile=False
    SystemConfig.IdentfierInARFile=str(SystemConfig.IdentfierInARFile).strip()
    SystemConfig.IdentifierInARFile=SystemConfig.IdentfierInARFile
    if str(SystemConfig.IdentfierInARFile) in str(lineList):
        identifierFoundInArFile=True

        #now that identifier is found, if further step fails we can safely fail those
        stopProcessing=True
        print "Identifier found in AR File :",SystemConfig.IdentifierInARFile
        Report.WriteTestStep("Identifier found in AR File : {0}".format(SystemConfig.IdentifierInARFile),"Identifier : {0} should exist in AR File".format(SystemConfig.IdentifierInARFile),"Same as expected","Passed");

        if "non-migrated" not in SystemConfig.arFileLocation:
            # remove header
            lineList = lineList[1:]

            # remove trailer
            lineList = lineList[:-1]

        for eachLine in lineList:
            print('*'*10);
            print('Searching for identifier: {0}'.format(SystemConfig.IdentifierInARFile));
            print('Line: {0}'.format(eachLine))
            print('*'*10);
            if str(SystemConfig.IdentifierInARFile) in str(eachLine):
                identifierFoundInArFile=True
                Report.WriteTestStep("Found AR File with identifier : {0}. Filename : {1}".format(SystemConfig.IdentifierInARFile,arFile),"Should be able to find AR File","Found AR File","Passed")

                #in the same line, the amount shall also exist
                if str(SystemConfig.amountInARFile) in str(eachLine):
                    Report.WriteTestStep(
                                         "Amount :{0} should be present in fileName {1} having identifier {2}".format(
                                             SystemConfig.amountInARFile, arFile, SystemConfig.IdentifierInARFile),
                                         "Expected Amount should be present", "Expected Amount is present",
                                         "Passed")

                    return (True,stopProcessing)

        Report.WriteTestStep("Amount :{0} should be present in fileName {1} having identifier {2}".format(SystemConfig.amountInARFile,arFile,SystemConfig.IdentifierInARFile),"Expected Amount should be present","Expected Amount is not present","Failed")
    return (False,stopProcessing)

def validateArFile(rowNumber):
    global arFileFound
    global usFileFound

    arFileLocation = getCellValue(SystemConfig.sheet_obj, rowNumber, SystemConfig.ARFileColNumber)
    SystemConfig.arFileLocation = str(arFileLocation).strip()
    newFiles=[]
    if str(arFileLocation).strip() is None:
        return

    if (arFileLocation[-1] == "/"):
        arFileLocation = arFileLocation[:-1]
    try:
        retries = 0
        while len(newFiles) == 0 and retries != 6:
            newFiles = sftpCustom.getListOfFilesAfterSpecificTimestamp(arFileLocation,lastDumpTime.lastDumpTime)
            time.sleep(10)
            retries += 1

    except Exception,e:
        Report.WriteTestStep("Path navigation failed","The given path [{0}] should exist on the server".format(arFileLocation),"AR file Location is either invalid or does not exist on the server","Failed")
        return False

    if len(newFiles)==0:
        #Write Report and return
        Report.WriteTestStep("Could not find any new AR files","AR files should have been generated after the file was dumped at timestamp : {0} ".format(lastDumpTime.lastDumpTime),"New AR files were not generated","Failed")
        return False

    SystemConfig.arFileTempDir="DownloadedFiles\\ARFiles\\"+str(time.time())

    os.makedirs(SystemConfig.arFileTempDir)

    sourceFileContents=SystemConfig.sourceFileCurrentContent

    for files in newFiles:
        #download each file to the same folder
        sftpCustom.downloadFile(arFileLocation+r"/"+files,SystemConfig.arFileTempDir+"\\"+files)
        (res,stopProcessing)=checkArFileContent(SystemConfig.arFileTempDir+"\\"+files)

        if stopProcessing:
            break

    if not res:
        if not stopProcessing:
            #this actually means that AR identifier was found, but amount did not match
            Report.WriteTestStep("Identifier not found in AR file : {0}".format(SystemConfig.IdentifierInARFile), "Identifier was expected", "Identifier not found","Failed")

        Report.WriteTestStep("Validation for AR File failed",
                                "Validation should be successful",
                                "Validation failed. Following files were scanned : {0}".format(newFiles), "Failed")
    else:
        Report.WriteTestStep("Validation of AR file successful","Validation should pass","Same as expected","Passed")
    return res

def validateUsFile(rowNumber):
    global usFileFound
    usFileLocation = getCellValue(SystemConfig.sheet_obj, rowNumber, SystemConfig.USFileColNumber)

    if (usFileLocation[-1] == "/"):
        usFileLocation = usFileLocation[:-1]

    newFiles = []
    retries = 0
    while len(newFiles) == 0 and retries != 6:
        newFiles = sftpCustom.getListOfFilesAfterSpecificTimestamp(usFileLocation,lastDumpTime.lastDumpTime)
        time.sleep(10)
        retries += 1

    print "New US Files detected : ",newFiles

    if len(newFiles)==0:
        #Write Report and return
        Report.WriteTestStep("Could not find any new US files","US file should have been generated","New US files were not generated","Failed")
        return False

    SystemConfig.usFileTempDir="DownloadedFiles\\USFiles\\"+str(time.time())
    os.makedirs(SystemConfig.usFileTempDir)
    sourceFileContents=SystemConfig.sourceFileCurrentContent

    #get the current Depository Account Number
    sourceFileContents=sourceFileContents.replace(str(SystemConfig.depositAccountCurrentNumber), str(SystemConfig.USFileDepositoryAccountCurrentNumber))

    for files in newFiles:
        #download each file to the same folder
        sftpCustom.downloadFile(usFileLocation+r"/"+files,SystemConfig.usFileTempDir+"\\"+files)
        if checkUsFileContents(sourceFileContents,SystemConfig.usFileTempDir+"\\"+files):
            Report.WriteTestStep("Validation for US File passed","Contents of raw file should be present in US File : {0}.\nRaw file content is : [{1}]".format(files,sourceFileContents),"Contents of raw file are present in US File","Passed")
            return True

    Report.WriteTestStep("Validation for US File failed",
                         "Contents of raw file should be present in one of the US Files : {0}.\nRaw file content is : [{1}]".format(newFiles,sourceFileContents),
                         "Contents of raw file are NOT present in any of the US Files", "Failed")
    return False

def checkUsFileContents(stringToMatch,FilenameToCheck):
    global usFileFound
    fileContents=readFileContentsAsString(FilenameToCheck)
    if stringToMatch in fileContents:
        return True
    return False

def run(excelFile):
    global arFileFound
    global usFileFound
    (wb_obj,sheet_obj,maxRows,maxColumns)=basicOperations(excelFile)
    SystemConfig.wb_obj=wb_obj
    SystemConfig.sheet_obj=sheet_obj
    SystemConfig.maxRows=maxRows
    SystemConfig.maxCol=maxColumns
    SystemConfig.sourceFileName=getFileNameFromExcel(sheet_obj,maxRows,maxColumns)
    SystemConfig.zipFile=getZipFilenameFromExcel(sheet_obj,maxRows,maxColumns)
    setColumnNumbersForFileValidations()

    Report.InitializeReporting(excelFile)
    if(not sftpCustom.Connect()):
        Report.WriteTestStep("Validate connectivity with the server",
                             "Should be able to be connected to the server",
                             "Unable to connected to the server", "Failed")
        Report.GeneratePDFReport()
        sys.exit(-1)

    Col=9
    currentRow=Config.rowNumberWhereFieldsArePlaced+1
    SystemConfig.startingRowNumberForRecordProcessing=currentRow

    SystemConfig.sourceFileEntireContent=readFileContentsAsList("FilesToUpload\\"+SystemConfig.sourceFileName)

    loopCtr=Config.rowNumberWhereFieldsArePlaced;
    #loopCtr=0;
    while True:

        """process each row, when empty row is encountered break"""

        loopCtr+=1
        currentRow=loopCtr;
        arFileFound=False
        usFileFound=False

        print('Max rows : ',maxRows)
        if loopCtr>=maxRows:
            break

        print "Loop count is : {0}".format(loopCtr);
        #time.sleep(1)

        SystemConfig.sourceFileCurrentContent = SystemConfig.sourceFileEntireContent[loopCtr-Config.rowNumberWhereFieldsArePlaced-1].strip("\n")
        print('Source file current content : {0}'.format(SystemConfig.sourceFileCurrentContent));
        #time.sleep(10)
        accountNumber = SystemConfig.sourceFileCurrentContent[48:68]

        if  getCellValue(SystemConfig.sheet_obj,currentRow,SystemConfig.RawFileColNumber) is not None:
            Report.WriteTestCase("Row# : {0}".format(loopCtr), "Raw File Validation for Account: {0}".format(accountNumber))
            validateRawFile(currentRow)
            Report.evaluateIfTestCaseIsPassOrFail()

        if  getCellValue(SystemConfig.sheet_obj, currentRow, SystemConfig.RenameFileColNumber) is not None:
            Report.WriteTestCase("Row# : {0}".format(loopCtr), "Rename File Validation for Account: {0}".format(accountNumber))
            validateRenameFile(currentRow)
            Report.evaluateIfTestCaseIsPassOrFail()

        if  getCellValue(SystemConfig.sheet_obj, currentRow, SystemConfig.ErrorFileColNumber) is not None:
            print "Cell value for error file is : [{0}] at row,col : {1},{2}".format(getCellValue(SystemConfig.sheet_obj, currentRow, SystemConfig.ErrorFileColNumber),currentRow,SystemConfig.ErrorFileColNumber)
            Report.WriteTestCase("Row# : {0}".format(loopCtr), "Error File Validation for Account: {0}".format(accountNumber))
            validateErrorFile(currentRow)
            Report.evaluateIfTestCaseIsPassOrFail()

        if  getCellValue(SystemConfig.sheet_obj, currentRow, SystemConfig.ARFileColNumber) is not None:
            SystemConfig.amountInARFile=getCellValue(SystemConfig.sheet_obj,currentRow,SystemConfig.ARAmountColNumber)
            SystemConfig.IdentifierInARFile=getCellValue(SystemConfig.sheet_obj,currentRow,SystemConfig.IdentifierInARFileColNumber)
            SystemConfig.IdentfierInARFile=SystemConfig.IdentifierInARFile
            print('*'*20)
            print('Row count : {0}'.format(currentRow))
            print('LoopCtr count : {0}'.format(loopCtr))
            print('Identifer in AR file : {0}'.format(SystemConfig.IdentifierInARFile));
            print('Amount in AR File : {0}'.format(SystemConfig.amountInARFile))
            print('*'*20)
            Report.WriteTestCase("Row# : {0}".format(loopCtr), "AR File Validation for Account: {0}".format(accountNumber))
            validateArFile(currentRow)
            Report.evaluateIfTestCaseIsPassOrFail()

        if  getCellValue(SystemConfig.sheet_obj, currentRow, SystemConfig.USFileColNumber) is not None:
            SystemConfig.USFileDepositoryAccountCurrentNumber=getCellValue(SystemConfig.sheet_obj,currentRow,SystemConfig.USFileDepositoryAccountColNumber)
            SystemConfig.depositAccountCurrentNumber=getCellValue(SystemConfig.sheet_obj,currentRow,SystemConfig.depositAccountColNumber)

            Report.WriteTestCase("Row# : {0}".format(loopCtr), "US File Validation for Account: {0}".format(accountNumber))
            validateUsFile(currentRow)
            Report.evaluateIfTestCaseIsPassOrFail()
        #currentRow+=1

    Report.GeneratePDFReport()
    sftpCustom.closeConnection()

if __name__ == '__main__':
    try:
        run(os.path.join(Config.assetsPath, "US Innove.xlsx"))
    except Exception,e:
        traceback.print_exc()

    Report.GeneratePDFReport()
    sftpCustom.closeConnection()